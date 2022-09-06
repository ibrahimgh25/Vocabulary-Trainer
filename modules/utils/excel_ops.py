import pandas as pd
from typing import Optional, Iterable, Union

import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Side, Border, Font, Alignment
from openpyxl.styles import PatternFill, NamedStyle


def get_excel_df(excel_file:str, sheetname:Optional[str]=None) -> pd.DataFrame:
    """
    Reads an Excel sheet with formulas and return the dataframe
    :param excel_file: path to Excel file
    :param sheetname: name of the sheet to read (default: None)
    :returns: dataframe found in the Excel sheet

    """
    wb = load_workbook(excel_file)
    if sheetname is None:
        sheet = wb.active
    else:
        sheet = wb[sheetname]
    df = pd.DataFrame(sheet.values)
    # Get the column names from the first row
    df.columns = df.iloc[0]
    # Delete the first row
    df = df.iloc[1:].fillna('')
    return df.reset_index(drop=True)

def style_dict(header_style:str, body_style:str) -> dict[str, str]:
    """
    Just returns a dictionary with the entered_styles
    :param header_style: the name of the header style
    :param body_style: the name of the body style
    :returns: a dictionary with the header and body styles

    """
    # Just make sure that the input are strings, not some random datatype
    assert type(header_style) == str, 'header_style should be a string'
    assert type(body_style) == str, 'body_style should be a string'
    return {'Header':header_style, 'Body':body_style}

def find_style(style_name:str) -> Union[NamedStyle, str]:
    """
    Returns a cell style based on the name
    :param style_name: the name of the needed style
    :returns: an openpyxl.styles.NamedStyle configured according to the name, if the name doesn't exist, it returns the 
     name itself

    """
    dark_blue = '00002060'
    black = '00000000'
    white = '00FFFFFF'
    pure_blue = '000000FF'
    # I know this is a primitive way of doing it, but it's easy
    style = NamedStyle(name=style_name)
    if style_name == 'BlueHeaderCentered':
        # This style has all borders white and thick, dark blue fill, white-bold font
        # and its alignment is center
        style.font = Font(name='Calibri', bold=True, size=11, color=white)
        style.border = all_borders('thick', white)
        style.alignment = Alignment(horizontal='center', vertical='center')
        style.fill = PatternFill(fill_type='solid', start_color=dark_blue)
        return style
    
    if style_name == 'BodyCentered':
        # This style has all borders black and thing, no fill, black font, and its 
        # alignment is center
        style.font = Font(name='Calibri', bold=False, size=11, color=black)
        style.border = all_borders('thin', black)
        style.alignment = Alignment(horizontal='center', vertical='center')
        return style
    
    if style_name == 'BodyLeft':
        # This style is the same as the 'BodyCentered' but with text left aligned
        style.font = Font(name='Calibri', bold=False, size=11, color=black)
        style.border = all_borders('thin', black)
        style.alignment = Alignment(horizontal='left', vertical='center')
        return style

    if style_name == 'CustomHyperlink':
        style.border = all_borders('thin', black)
        style.font = Font(name='Calibri', u='single', size=11, color=pure_blue)
        style.alignment = Alignment(horizontal='center', vertical='center')
        return style
    else:
        # This returns style name if the name isn't covered in this function, it should
        # be a named style built-in in openpyxl
        return style_name

def add_styles(wb:Workbook)->Workbook:
    """
    Adds two custom styles to a workbook
    :param wb: the workbook to add the styles to
    :returns: the workbook with the styles added

    """
    styles = ['BlueHeaderCentered', 'BodyCentered', 'BodyLeft', 'CustomHyperlink']
    for style in styles:
        try:
            wb.add_named_style(find_style(style))
        except ValueError:
            # This is to avoid raising errors when the style already exists
            pass
    return wb

def all_borders(line_style:NamedStyle, color:str='00000000') -> Border:
    """
    Returns a border with all sides at the specified type and color
    :param line_style: the style of the border
    :param color: the HEX color of the border  (Default black)

    """
    side = Side(style=line_style, color=color)
    border = Border(left=side, top=side, right=side, bottom=side)
    return border
    
def adjust_column_width(ws:Worksheet, col_index:int, width:int) -> Worksheet:
    """
    Adjusts the widths of a column in a worksheet
    :param ws: the worksheet containing the column
    :param col_index: the index of the column
    :param width: the required width for the column
    :returns: the worksheet after editing the column
    """
    assert width >= 0, "A negative width isn't allowed"
    
    letter = get_column_letter(col_index + 1)
    if width == 0:
        ws.column_dimensions[letter].autosize = True
    else:
        ws.column_dimensions[letter].width = width
    return ws

def df_to_excel(df:pd.DataFrame, filename:str, sheet_name:str, styles:Iterable[dict]=None, dims:Iterable[int]=None):
    """
    Saves a dataframe in an Excel sheet with specified styles
    :param df: the dataframe to save
    :param filename: path to the Excel file
    :param sheet_name: the sheet name
    :param styles: list containing the styles to be applied for every column
    :param dims: the width of each column in Excel
    Note: if the function fails, it'll give you an opportunity to close the file and retry again, this helped
     me a lot when testing, so I am leaving it

    """
    # Try to load the existing workbook into a file
    try:
        wb = load_workbook(filename)
        existing_sheets = wb.sheetnames
        if sheet_name in existing_sheets:
            # Make sure to create the function at the same index
            idx = wb.worksheets.index(wb[sheet_name])
            wb.remove(wb[sheet_name])
            ws = wb.create_sheet(title=sheet_name, index=idx)
        else:
            ws = wb.create_sheet(title=sheet_name)
    # If the file doesn't exist, create a new workbook
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    # Hide the gridlines from the Excel sheet
    ws.sheet_view.showGridLines = False
    # Copy the dataframe rows to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb = add_styles(wb)
    len_match_err = "The number of {} should match the number of columns in the dataframe"
    # Apply styling if a styles' list were passed to the function
    if styles:
        # Make sure each column is assigned a style
        assert len(styles) == len(df.columns), len_match_err.format("styles")
        # Iterate over each column and apply styling separately
        for col_nb, col in enumerate(ws.iter_cols()):
            for row_nb, cell in enumerate(col):
                # If row_nb==0 then we're at the header
                if row_nb==0:
                    cell.style = styles[col_nb]['Header']
                else:
                    cell.style = styles[col_nb]['Body']
    # Apply column widths if they were given
    if dims:
        # Make sure each column is assigned a width
        assert len(dims) == len(df.columns), len_match_err.format("assigned widths")
        # Iterate over columns and apply width to each one
        for index, _ in enumerate(ws.iter_cols()):
            ws = adjust_column_width(ws, index, dims[index])
    # The following block is for saving the workbook, the try block is only for when the file 
    # we're trying to save to is open, this happens a lot to me in testing, so I added the try
    # block
    try:
        wb.save(filename)
    except PermissionError:
        print(filename, " may be still open, please close and press any key to try again")
        os.system("pause")
        wb.save(filename)


def save_to_excel(df:pd.DataFrame, filename:str, sheetname:str) -> None:
    """
    Saves the language dataframe to an Excel file
    :param df: the dataframe to save
    :param filename: the path of hte Excel file to save under 
    :param sheetname: the sheet name of the saved database
    """
    # Create a list of dictionaries with the styles for each column
    styles = [{'Header':'BlueHeaderCentered', 'Body':'BodyLeft'}] * len(df.columns)
    # Create a list of widths for each column
    # Create a list of widths with length of 10
    dims = [0] * len(df.columns)
    # Save the dataframe to an Excel file
    df_to_excel(df, filename, sheetname, styles, dims)