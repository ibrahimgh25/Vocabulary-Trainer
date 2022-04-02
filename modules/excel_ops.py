import pandas as pd
from typing import Optional, Iterable, Union

import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Side, Border, Font, Alignment
from openpyxl.styles import PatternFill, NamedStyle


def get_excel_df(excel_file:str, sheetname:Optional[str]=None) -> pd.DataFrame:
    """
    Reads an excel sheet with formulas and return the dataframe.
    :param excel_file: path to excel file
    :param sheetname: name of the sheet to read (default: None)
    :return: dataframe found in the excel sheet
    """
    wb = load_workbook(excel_file)
    if sheetname is None:
        sheet = wb.active
    else:
        sheet = wb[sheetname]
    df = pd.DataFrame(sheet.values)
    # Chaneg the columns names to the first row
    df.columns = df.iloc[0]
    # Delete the first row
    df = df.iloc[1:].fillna('')
    return df.reset_index(drop=True)

def style_dict(header_style:str, body_style:str) -> dict[str, str]:
    '''Just returns a dictionary with the entered_styles
        :params header_style: the name of the header style
        :params body_style: the name of the body style
        :returns: a dictionary with the header and body styles
    '''
    # Just make sure that the input are strings, not some random datatype
    assert type(header_style) == str, 'header_style should be a string'
    assert type(body_style) == str, 'body_style should be a string'
    return {'Header':header_style, 'Body':body_style}

def find_style(style_name:str) -> Union[NamedStyle, str]:
    ''' Returns a cell style based on the name
        :params style_name: the name of the needed style
        :returns: an openpyxl.styles.NamedStyle configured according to the name, if the name doesn't exist,
         it returns the name itself
    '''
    dark_blue = '00002060'
    black = '00000000'
    white = '00FFFFFF'
    pure_blue = '000000FF'
    # I know this is a premitive way of doing it, but it's easy
    style = NamedStyle(name=style_name)
    if style_name == 'BlueHeaderCentered':
        # This style has all borders white and thick, dark blue fill, white-bold font
        # and its alignment is center
        style.font = Font(name='Calibri', bold=True, size=11, color=white)
        style.border = all_borders('thick', white)
        style.alignment = Alignment(horizontal='center', vertical='center')
        style.fill = PatternFill(fill_type='solid', start_color=dark_blue)
        return style
    
    if style_name == 'BodyCentered':
        # This style has all borders black and thing, no fill, black font, and its 
        # alignment is center
        style.font = Font(name='Calibri', bold=False, size=11, color=black)
        style.border = all_borders('thin', black)
        style.alignment = Alignment(horizontal='center', vertical='center')
        return style
    
    if style_name == 'BodyLeft':
        # This style is the same as the 'BodyCentered' but with text left aligned
        style.font = Font(name='Calibri', bold=False, size=11, color=black)
        style.border = all_borders('thin', black)
        style.alignment = Alignment(horizontal='left', vertical='center')
        return style

    if style_name == 'CustomHyperlink':
        style.border = all_borders('thin', black)
        style.font = Font(name='Calibri', u='single', size=11, color=pure_blue)
        style.alignment = Alignment(horizontal='center', vertical='center')
        return style
    else:
        # This returns style name if the name isn't covered in this function, it should
        # be a named style built-in in openpyxl
        return style_name

def add_styles(wb:Workbook)->Workbook:
    '''Adds two custom styles to a workbook
    Parameters:
        :params wb: the workbook to add the styles to
        :returns: the workbook with the styles added
    '''
    styles = ['BlueHeaderCentered', 'BodyCentered', 'BodyLeft', 'CustomHyperlink']
    for style in styles:
        try:
            wb.add_named_style(find_style(style))
        except ValueError:
            # This is to avoid raising errors when the style already exists
            pass
    return wb

def all_borders(line_style:NamedStyle, color:str='00000000') -> Border:
    ''' Returns a border with all sides at the specified type and color
    Parameters:
        color (string): the hex code of the color, the default is black color
        line_style (string): must be one of the types allowed with openpyxl.styles.Side()
    Returns openpyxl.styles.Border()
    '''
    side = Side(style=line_style, color=color)
    border = Border(left=side, top=side, right=side, bottom=side)
    return border
    
def adjust_column_width(ws:Worksheet, col_index:int, width:int) -> Worksheet:
    ''' Adjusts the widths of the a column in worksheet
    Parameters:
        ws (openpyxl worksheet): the excel worksheet
        col_index (int): the index of the column to adjust
        width (double): the width we want to assign (0 means autofit), no negative numbers
         are allowed
    Returns the worksheet with the column width adjusted
    '''
    assert width >= 0, "A negative width isn't allowed"
    
    letter = get_column_letter(col_index + 1)
    if width == 0:
        ws.column_dimensions[letter].autosize = True
    else:
        ws.column_dimensions[letter].width = width
    return ws

def df_to_excel(df:pd.DataFrame, filename:str, sheet_name:str, styles:Iterable[dict]=None, dims:Iterable[int]=None):
    '''Saves a dataframe in an excel sheet with specified styles
    Parameters:
        df (pandas dataframe): the dataframe we need to save
        filename (string): the name (path) of the excel file
        sheet_name (string): the name of the sheet (originally I allowed for the sheet name to be 
         set to None, which uses default naming, but because of some problems that complicates the code
          I decided not to, so you have to give a name to your sheet when using this function
        styles (list of dictionaries): each entry contains the style of a specific column,
         each element should be a dictionary with the keys 'Header' and 'Body', each one linking 
         to a name of a style (string)
        dims (list of integers): list containing the sequence of widths for the columns, it 
         should have the same length as the number of columns in df, a value of "0" means autofit
    '''
    # Try to load the exising workbook into a file
    try:
        wb = load_workbook(filename)
        existing_sheets = wb.sheetnames
        if sheet_name in existing_sheets:
            sheet_to_remove = wb.get_sheet_by_name(sheet_name)
            wb.remove(sheet_to_remove)
            ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb.create_sheet(title=sheet_name)
    # If the file doesn't exist, create a new workbook
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    # Hide the gridlines from the excel sheet
    ws.sheet_view.showGridLines = False
    # Copy the dataframe rows to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb = add_styles(wb)
    len_match_err = "The number of {} should match the number of columns in the dataframe"
    # Apply styling if a styles' list were passed to the function
    if styles:
        # Make sure each column is assigned a style
        assert len(styles) == len(df.columns), len_match_err.format("styles")
        # Iterate over each column and apply styling seperatly
        for col_nb, col in enumerate(ws.iter_cols()):
            for row_nb, cell in enumerate(col):
                # If row_nb==0 then we're at the header
                if row_nb==0:
                    cell.style = styles[col_nb]['Header']
                else:
                    cell.style = styles[col_nb]['Body']
    # Apply column widths if they where given
    if dims:
        # Make sure each column is assigned a width
        assert len(dims) == len(df.columns), len_match_err.format("assigned widths")
        # Iterate over columns and apply width to each one
        for index, _ in enumerate(ws.iter_cols()):
            ws = adjust_column_width(ws, index, dims[index])
    # The following block is for saving the workbook, the try block is only for when the file 
    # we're trying to save to is open, this happens a lot to me in testing so I added the try
    # block
    try:
        wb.save(filename)
    except PermissionError as e:
        print(filename, " may be still open, please close and press any key to try again")
        os.system("pause")
        wb.save(filename)


def save_to_excel(language_df:pd.DataFrame, filename:str, sheetname:str):
    '''Saves the language dataframe to an excel file
    Parameters:
        :params language_df: the dataframe containing the language data
        :params sheetname: the name of the sheet to save the dataframe to
    '''
    # Create a list of dictionaries with the styles for each column
    styles = [{'Header':'BlueHeaderCentered', 'Body':'BodyLeft'}] * len(language_df.columns)
    # Create a list of widths for each column
    # Create a list of widths with length of 10
    dims = [0] * len(language_df.columns)
    # Save the dataframe to an excel file
    df_to_excel(language_df, filename, sheetname, styles, dims)
    

